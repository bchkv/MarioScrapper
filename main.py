import logging
import re
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import cursor
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import column_index_from_string
import os
import atexit
import shutil

# To install all requirements enter the command:
# pip install -r requirements.txt

# https://stackoverflow.com/questions/46419607/how-to-automatically-install-required-packages-from-a-python-script-as-necessary#:~:text=You%20can%20use%20pipreqs%20to,pip%20install%20pipreqs%20pipreqs%20.

logging.basicConfig(filename='log_file.txt',
                    filemode='w',
                    format='%(message)s',
                    level='DEBUG')
headers = {'User-Agent': 'Mozilla/5.0'}

login_url = 'http://www.edumich.gob.mx/sigem_tel/index/1/'
page_with_tables = "http://www.edumich.gob.mx/sigem_tel/sisat_registro_2223/" \
                   "3c2a3acc3ebccb7e62352756b14fd812b6913fe1/I2223/"

Path(f"tables").mkdir(parents=True, exist_ok=True)

cursor.hide()
atexit.register(cursor.show)

# For each xlsx file we need to memorise: school, file name, course year, school group
tables_data = dict()
faulty_tables = list()


class School:
    table_count = int(0)
    table_count_check = int(0)

    def __init__(self, credentials):
        self.login, self.password = credentials.split(':')
        self.name = self.login.split('_')[0]

    def get_tables(self):
        with requests.session() as session:
            response = session.post(login_url, data={'inputEmail': '',
                                                     'inputPassword': self.password, 'grabar': 'si'})
            response.raise_for_status()
            if response.url == login_url:
                print(f"Could not login to the school {self.login} with the password '{self.password}'!")
                print("Check credentials for that school and start again!")
                exit(1)

            response = session.get(page_with_tables, headers=headers)
            response.raise_for_status()

            # Now we need to get urls of tables, they always have titles "Concentrado de Información"
            soup = BeautifulSoup(response.content, 'html.parser')
            number_of_tables = len(soup.find_all(title='Concentrado de Información'))
            print(f"Found {number_of_tables} tables for the school {self.name}.")
            School.table_count += len(soup.find_all(title='Concentrado de Información'))
            for count, table_page_element in enumerate(soup.find_all(title="Concentrado de Información")):
                table_url = table_page_element.get('href')

                table_page = session.get(table_url, headers=headers)
                table_page.raise_for_status()

                table_page_soup = BeautifulSoup(table_page.content, 'html.parser')
                download_table_url = table_page_soup.find(class_="btn btn-warning btn-sm").get('href')
                print(f"Downloading table {count + 1} of {number_of_tables}...", end='')
                print('\r', end='')
                table_file = session.get(download_table_url, allow_redirects=True, stream=True)

                # If the file is not an Excel-table, show error message and continue with the next file
                correct_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                if table_file.headers['Content-Type'] != correct_content_type:
                    print(f'''Something wrong with the file!\n
                              Log in manually in the school {self.name} and then check the link: {download_table_url}\n
                              You can download this file manually and put it in the folder 'Tables''')
                    file_name = f"Unknown table {count + 1} from {self.name}"
                    faulty_tables.append(file_name)
                    tables_data[file_name] = {'school': self.name, 'course': None, 'group': None,
                                              'download_url': download_table_url}
                    continue

                file_name = re.search(f'^.*="(.+)"$', table_file.headers['Content-Disposition']).group(1)
                with open(f"tables/{self.name}_{file_name}", 'wb') as file:
                    file.write(table_file.content)
                    School.table_count_check += 1

                class_data = re.search(r'"(\d)([a-zA-Z])"', file_name)
                if class_data:
                    course_year = class_data.group(1)
                    group = class_data.group(2)
                else:
                    course_year = None
                    group = None
                    faulty_tables.append(file_name)

                tables_data[file_name] = {'school': self.name, 'course': course_year, 'group': group,
                                          'download_url': download_table_url}

                # check_table(f"{self.name}_{file_name}")


def check_table(table_name):
    try:
        pd.read_excel("tables/" + table_name)
    except ValueError:
        print(f"The table {table_name} seems to be faulty!")
        faulty_tables.append(table_name)
        return False
    return True


def proceed_tables(_min, _max):
    """
    Function checks if tables in the folder are not faulty
    Also it categorises them by school name, year and class letter and stores it into a dictionary
    Also it colores cells in the row Average whith values within the range [_min, _max]
    """

    tables_dict = dict()
    found_tables_in_folder = int(0)
    recorded_tables_in_dict = int(0)

    tables_names_list = list(map(str, Path("tables").iterdir()))
    tables_names_list = list(map(lambda x: x.replace("tables/", ""), tables_names_list))

    tables_names_list.sort()
    for table_name in tables_names_list:
        if table_name == ".DS_Store":
            continue
        # check if file name is *.xlsx format
        if not re.search(r"^.*\.xlsx$", table_name):
            continue
        found_tables_in_folder += 1
        table_name = table_name.replace("tables/", "")
        # If the table is not faulty
        if check_table(table_name):
            # We got to get school name, year and letter with regex
            school_name = re.search(r"(^[0-9a-zA-Z]*)_", table_name).group(1)
            year = re.search(r'"([0-9])[A-Za-z]"', table_name).group(1)
            letter = re.search(r'"[0-9]([A-Za-z])"', table_name).group(1)
            # Write to the dict with corresponding keys
            if school_name not in tables_dict:
                tables_dict[school_name] = dict()
            if year not in tables_dict[school_name]:
                tables_dict[school_name][year] = dict()
            tables_dict[school_name][year][letter] = table_name

        else:
            if 'faulty_tables' not in tables_dict:
                tables_dict['faulty_tables'] = list()
            tables_dict['faulty_tables'].append(table_name)

    # We need to calculate number of schools recorded in the dictionary
    for s in tables_dict:
        if s != 'faulty_tables':
            for y in tables_dict[s]:
                recorded_tables_in_dict += len(tables_dict[s][y])
    recorded_tables_in_dict += len(tables_dict['faulty_tables'])

    # pp.pprint(tables_dict)
    print(f"Recorded {recorded_tables_in_dict} of {found_tables_in_folder} tables")

    red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')

    # Cleaning out the folder
    shutil.rmtree('out', ignore_errors=True)

    print()

    for table_name in tables_names_list:
        if not re.search(r"^.*\.xlsx$", table_name) or table_name in tables_dict['faulty_tables']:
            continue

        print(f"Working on {table_name}...", end='')
        print('\r', end='')

        workbook = load_workbook(f"tables/{table_name}")
        worksheet = workbook.active

        # Counting number of pupils
        pupils_count = 0
        row = 13
        while True:
            if not isinstance(worksheet.cell(row, column_index_from_string('A')).value, int):
                break
            pupils_count += 1
            row += 1

        worksheet.conditional_formatting.add(f'Y13:Y{13+pupils_count-1}', CellIsRule(operator='between', formula=[str(_min), str(_max)],
                                                                   stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'AN13:AN{13+pupils_count-1}', CellIsRule(operator='between', formula=[str(_min), str(_max)],
                                                                   stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'BC13:BC{13+pupils_count-1}', CellIsRule(operator='between', formula=[str(_min), str(_max)],
                                                                   stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'BR13:BR{13+pupils_count-1}', CellIsRule(operator='between', formula=[str(_min), str(_max)],
                                                                   stopIfTrue=True, fill=red_fill))

        for x in range(13, pupils_count + 13):
            if worksheet.cell(x, column_index_from_string('Y')).value and _min <= worksheet.cell(x, column_index_from_string('Y')).value <= _max:
                cell = worksheet[f"Y{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('AN')).value and _min <= worksheet.cell(x, column_index_from_string('AN')).value <= _max:
                cell = worksheet[f"AN{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BC')).value and _min <= worksheet.cell(x, column_index_from_string('BC')).value <= _max:
                cell = worksheet[f"BC{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BR')).value and _min <= worksheet.cell(x, column_index_from_string('BR')).value <= _max:
                cell = worksheet[f"BR{x}"]
                cell.fill = red_fill

            if worksheet.cell(x, column_index_from_string('AB')).value not in [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"AB{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('AQ')).value not in [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"AQ{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BF')).value not in [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"BF{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BU')).value not in [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"BU{x}"]
                cell.fill = red_fill

            if not worksheet.cell(x, column_index_from_string('Z')).value == "SI" and\
                    worksheet.cell(x, column_index_from_string('AA')).value != None:
                cell = worksheet[f"Z{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('AO')).value == "SI" and\
                    worksheet.cell(x, column_index_from_string('AP')).value != None:
                cell = worksheet[f"AO{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('BD')).value == "SI" and\
                    worksheet.cell(x, column_index_from_string('BE')).value != None:
                cell = worksheet[f"BD{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('BS')).value == "SI" and\
                    worksheet.cell(x, column_index_from_string('BT')).value != None:
                cell = worksheet[f"BS{x}"]
                cell.fill = red_fill

        Path(f"out").mkdir(parents=True, exist_ok=True)
        workbook.save(f"out/test_{table_name}")

    print("Done!")
    # NOTE: Deprecated
    # for school in tables_dict:
    #     out_path = f"/out/{school}.xlsx"
    #     book = load_workbook(out_path)
    #     writer = pd.ExcelWriter(out_path, engine='openpyxl')
    #     writer.book = book
    #     if school != 'faulty_tables':
    #         for year in tables_dict[school]:
    #             school_dataframes = list()
    #             year_data = None
    #             for letter in tables_dict[school][year]:
    #                 letter_data = get_dataframes(tables_dict[school][year][letter])
    #                 if not year_data:
    #                     year_data = letter_data
    #                 else:
    #                     for x in range(0,4):
    #                         year_data[x].append(letter_data[x], ignore_index=False)
    #
    #             school_dataframes.append(year_data)


def get_dataframes(file_name):
    """Append data from list of tables to one single table using Pandas
       Function receives table's names as arguments
       Returns tuple with 4 dataframes for each period"""

    file_name = 'test/16ETV0193A_2022_2023_"1A".xlsx'
    file2 = 'test/16ETV0193A_2022_2023_"1B".xlsx'
    file3 = 'test/16ETV0193A_2022_2023_"1C".xlsx'
    file_x = 'tables/16ETV0430M_2022_2023_"3A".xlsx'
    
    # change all cupr_x parts' data type to str, then concatenate it into single column cupr
    cupr_df = pd.read_excel(file_name, skipfooter=13, usecols="B:K", header=11, names=['cupr_1', 'cupr_2', 'cupr_3', 'cupr_4', 'cupr_5', 'cupr_6', 'cupr_7', 'cupr_8', 'cupr_9', 'name'], dtype={'C': np.int32, })
    cupr_df = cupr_df.reset_index(drop=True)
    # print(cupr_df.head(5))
    # Concatenating cupr parts into single string
    cupr_df["cupr"] = cupr_df["cupr_1"].astype(str) + cupr_df["cupr_2"].astype(str) + cupr_df["cupr_3"].astype(str) \
                      + cupr_df["cupr_4"].astype(str) + cupr_df["cupr_5"].astype(str) + cupr_df["cupr_6"].astype(str) \
                      + cupr_df["cupr_7"].astype(str) + cupr_df["cupr_8"].astype(str) + cupr_df["cupr_9"].astype(str)
    cupr_df = cupr_df.drop(columns=['cupr_1', 'cupr_2', 'cupr_3', 'cupr_4', 'cupr_5',
                                    'cupr_6', 'cupr_7', 'cupr_8', 'cupr_9'])
    # IDEA: Implement letter-auto-detection
    cupr_df["letter"] = 'A'

    # Data for the first period
    scores_df_1 = pd.read_excel(file_name, usecols="O:X", header=6, skipfooter=13)
    scores_df_1 = scores_df_1.drop(range(0, 5))
    scores_df_1 = scores_df_1.reset_index(drop=True)
    
    scores_df_1.insert(loc=0, column='period', value='1')

    sum_df_1 = pd.read_excel(file_name, usecols='Y', header=5, skipfooter=13, names=['average'])
    sum_df_1 = sum_df_1.drop(range(0, 6))
    sum_df_1 = sum_df_1.reset_index(drop=True)

    # Data for the second period
    scores_df_2 = pd.read_excel(file_name, usecols="AD:AM", header=6, skipfooter=13)
    scores_df_2 = scores_df_2.drop(range(0, 5))
    scores_df_2 = scores_df_2.reset_index(drop=True)

    scores_df_2.insert(loc=0, column='period', value='2')

    sum_df_2 = pd.read_excel(file_name, usecols='AN', header=5, skipfooter=13, names=['average'])
    sum_df_2 = sum_df_2.drop(range(0, 6))
    sum_df_2 = sum_df_2.reset_index(drop=True)

    # Data for the third period
    scores_df_3 = pd.read_excel(file_name, usecols="AS:BB", header=6, skipfooter=13)
    scores_df_3 = scores_df_3.drop(range(0, 5))
    scores_df_3 = scores_df_3.reset_index(drop=True)

    scores_df_3.insert(loc=0, column='period', value='3')

    sum_df_3 = pd.read_excel(file_name, usecols='BC', header=5, skipfooter=13, names=['average'])
    sum_df_3 = sum_df_3.drop(range(0, 6))
    sum_df_3 = sum_df_3.reset_index(drop=True)

    # Data for the final period
    scores_df_final = pd.read_excel(file_name, usecols="BH:BQ", header=6, skipfooter=13)
    scores_df_final = scores_df_final.drop(range(0, 5))
    scores_df_final = scores_df_final.reset_index(drop=True)

    scores_df_final.insert(loc=0, column='period', value="final")

    sum_df_final = pd.read_excel(file_name, usecols='BR', header=5, skipfooter=13, names=['average'])
    sum_df_final = sum_df_final.drop(range(0, 6))
    sum_df_final = sum_df_final.reset_index(drop=True)

    # (deprecated) add Approbas and Alumno Regular columns

    result_1 = pd.concat([cupr_df, scores_df_1, sum_df_1], axis="columns")
    result_2 = pd.concat([cupr_df, scores_df_2, sum_df_2], axis="columns")
    result_3 = pd.concat([cupr_df, scores_df_3, sum_df_3], axis="columns")
    result_final = pd.concat([cupr_df, scores_df_final, sum_df_final], axis="columns")

    # result_1.to_csv(path_or_buf="out.csv")
    # result_2.to_csv(path_or_buf="out_2.csv")

    return result_1, result_2, result_3, result_final


def download_tables():
    list_of_logins = list()
    with open("logins.txt", 'r') as file:
        for line in file:
            list_of_logins.append(line.strip())

    for login in list_of_logins:
        School(login).get_tables()

    print(f"\nFinished! Saved {School.table_count_check} of {School.table_count} tables.\n")

    for table in faulty_tables:
        print(f'Something went wrong with the table {table} from the school {tables_data[table]["school"]}.\n'
              f"Please, check it manually: log in into the school {tables_data[table]['school']} and visit "
              f"{tables_data[table]['download_url']} to download file\n")

    if faulty_tables:
        while True:
            print('You had a problem with one or more table files. You can resolve it manually or ignore it.\n'
                  'Print "y" once you are ready: ', end='')
            char = input()
            if char == 'y':
                break

# Remove while debugging
os.system('clear')
print("Press Ctrl + C to quit the program\n")
answer = input("Do you want to download tables (d) or just proceed them (p)? ")
if answer == 'd':
    download_tables()
elif answer not in ['d', 'p']:
    print("Wrong input!")
    exit(0)
min_inut, max_input = input("Enter min and max values to color row with average scores: ").split(' ')
print(f"Coloring values between {min_inut} and {max_input}...")
proceed_tables(int(min_inut), int(max_input))
