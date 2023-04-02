import pandas as pd
from pathlib import Path
from globals import faulty_tables
import re
from globals import tables_data
from schools import School
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import column_index_from_string
import shutil
import os


def categorize_tables(tables_names_list):
    """
    Categorizes tables by school name, year, and class letter.

    Args:
        tables_names_list (list): A list of table names.

    Returns:
        dict: A dictionary with categorized tables.
    """
    tables_dict = dict()
    tables_dict['faulty_tables'] = list()

    for table_name in tables_names_list:
        if table_name == ".DS_Store" or not re.search(r"^.*\.xlsx$", table_name):
            continue

        if check_table(table_name):
            school_name, year, letter = extract_table_info(table_name)
            add_table_to_dict(tables_dict, school_name, year, letter, table_name)
        else:
            add_faulty_table(tables_dict, table_name)

    return tables_dict


def extract_table_info(table_name):
    """
    Extracts school name, year, and class letter from a table name using regex.

    Args:
        table_name (str): The name of the table.

    Returns:
        tuple: A tuple containing school_name, year, and letter.
    """
    school_name = re.search(r"(^[0-9a-zA-Z]*)_", table_name).group(1)
    year = re.search(r'"([0-9])[A-Za-z]"', table_name).group(1)
    letter = re.search(r'"[0-9]([A-Za-z])"', table_name).group(1)
    return school_name, year, letter


def add_table_to_dict(tables_dict, school_name, year, letter, table_name):
    """
    Adds a table to the tables dictionary with the corresponding keys.

    Args:
        tables_dict (dict): The dictionary to add the table to.
        school_name (str): The name of the school.
        year (str): The year of the table.
        letter (str): The class letter of the table.
        table_name (str): The name of the table.

    Returns:
        None
    """
    if school_name not in tables_dict:
        tables_dict[school_name] = dict()
    if year not in tables_dict[school_name]:
        tables_dict[school_name][year] = dict()
    tables_dict[school_name][year][letter] = table_name


def add_faulty_table(tables_dict, table_name):
    """
    Adds a faulty table to the tables' dictionary.

    Args:
        tables_dict (dict): The dictionary to add the faulty table to.
        table_name (str): The name of the faulty table.

    Returns:
        None
    """
    if 'faulty_tables' not in tables_dict:
        tables_dict['faulty_tables'] = list()
    tables_dict['faulty_tables'].append(table_name)


def process_tables(_min, _max):
    """
    Processes tables in the folder and categorizes them by school name, year, and class letter.
    It also colors cells in the row 'Average' with values within the range [_min, _max].

    Args:
        _min (float): The minimum value of the range to color cells.
        _max (float): The maximum value of the range to color cells.

    Returns:
        None
    """

    tables_names_list = [str(table_path) for table_path in Path("tables").iterdir()]
    tables_names_list = [table_name.replace("tables/", "") for table_name in tables_names_list]
    tables_names_list.sort()

    tables_dict = categorize_tables(tables_names_list)

    out_directory = os.path.join(os.path.expanduser("~"), "Desktop/resultado")

    # Cleaning out the folder
    shutil.rmtree(out_directory, ignore_errors=True)

    red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')

    if 'faulty_tables' not in tables_dict:
        tables_dict['faulty_tables'] = list()
    for table_name in tables_names_list:
        if not re.search(r"^.*\.xlsx$", table_name) or table_name in tables_dict['faulty_tables']:
            continue

        print(f"Working on {table_name}...", end='')
        print('\r', end='')

        workbook = load_workbook(f"tables/{table_name}")
        worksheet = workbook.active

        # Counting number of pupils
        pupils_count = 0
        row = 13
        while True:
            if not isinstance(worksheet.cell(row, column_index_from_string('A')).value, int):
                break
            pupils_count += 1
            row += 1

        worksheet.conditional_formatting.add(f'Y13:Y{13+pupils_count-1}', CellIsRule(operator='between',
                                             formula=[str(_min), str(_max)], stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'AN13:AN{13+pupils_count-1}', CellIsRule(operator='between',
                                             formula=[str(_min), str(_max)], stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'BC13:BC{13+pupils_count-1}', CellIsRule(operator='between',
                                             formula=[str(_min), str(_max)], stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(f'BR13:BR{13+pupils_count-1}', CellIsRule(operator='between',
                                             formula=[str(_min), str(_max)], stopIfTrue=True, fill=red_fill))

        for x in range(13, pupils_count + 13):
            if worksheet.cell(x, column_index_from_string('Y')).value and\
                    _min <= worksheet.cell(x, column_index_from_string('Y')).value <= _max:
                cell = worksheet[f"Y{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('AN')).value and\
                    _min <= worksheet.cell(x, column_index_from_string('AN')).value <= _max:
                cell = worksheet[f"AN{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BC')).value and\
                    _min <= worksheet.cell(x, column_index_from_string('BC')).value <= _max:
                cell = worksheet[f"BC{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BR')).value and\
                    _min <= worksheet.cell(x, column_index_from_string('BR')).value <= _max:
                cell = worksheet[f"BR{x}"]
                cell.fill = red_fill

            if worksheet.cell(x, column_index_from_string('AB')).value not in\
                    [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"AB{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('AQ')).value not in\
                    [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"AQ{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BF')).value not in\
                    [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"BF{x}"]
                cell.fill = red_fill
            if worksheet.cell(x, column_index_from_string('BU')).value not in\
                    [None, "10/10", "9/9", "8/8", "7/7", "6/6", "5/5", "4/4", "3/3", "2/2", "1/1"]:
                cell = worksheet[f"BU{x}"]
                cell.fill = red_fill

            if not worksheet.cell(x, column_index_from_string('Z')).value == "SI" and\
                    not worksheet.cell(x, column_index_from_string('AA')).value:
                cell = worksheet[f"Z{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('AO')).value == "SI" and\
                    not worksheet.cell(x, column_index_from_string('AP')).value:
                cell = worksheet[f"AO{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('BD')).value == "SI" and\
                    not worksheet.cell(x, column_index_from_string('BE')).value:
                cell = worksheet[f"BD{x}"]
                cell.fill = red_fill
            if not worksheet.cell(x, column_index_from_string('BS')).value == "SI" and\
                    not worksheet.cell(x, column_index_from_string('BT')).value:
                cell = worksheet[f"BS{x}"]
                cell.fill = red_fill

        Path(out_directory).mkdir(parents=True, exist_ok=True)
        workbook.save(os.path.join(out_directory, table_name))


def download_tables():
    """
    Downloads tables from the website.

    Returns:
        None
    """
    list_of_logins = list()
    with open("data/logins.txt", 'r') as file:
        for line in file:
            list_of_logins.append(line.strip())

    for login in list_of_logins:
        School(login).get_tables()

    print(f"\n¡Terminado! Guardadas {School.table_count_check} de {School.table_count} tablas.\n")

    for table in faulty_tables:
        print(f'Something went wrong with the table {table} from the school {tables_data[table]["school"]}.\n'
              f"Please, check it manually: log in into the school {tables_data[table]['school']} and visit "
              f"{tables_data[table]['download_url']} to download file\n")


def check_table(table_name):
    """
    Checks if a table with the given table_name can be opened.

    Args:
        table_name (str): The name of the table to check.

    Returns:
        bool: True if the table can be opened, False otherwise.
    """
    try:
        pd.read_excel("tables/" + table_name)
    except ValueError:
        print(f"The table {table_name} seems to be faulty!")
        faulty_tables.append(table_name)
        return False
    return True
